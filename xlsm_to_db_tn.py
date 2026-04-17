#!/usr/bin/env python3
"""
청하 XLSM TN-DATA → 총질소_시험기록부
"""

import openpyxl
from datetime import datetime
import mysql.connector

DB_CONFIG = {
    'host': '1110s.synology.me',
    'user': 'eta_user',
    'password': '1212xx!!AA',
    'database': 'eta_db',
    'charset': 'utf8mb4'
}

xlsm_file = '/Users/azrael/Documents/ETA/Docs/CHUNGHA-김가린.xlsm'

try:
    print("XLSM 파일 로드 중...")
    wb = openpyxl.load_workbook(xlsm_file)
    ws = wb['TN-DATA']
    print(f"✓ TN-DATA 시트 로드 (max_row: {ws.max_row})")

    records = []

    # 모든 행 처리
    for row_num in range(2, ws.max_row + 1):
        date_val = ws.cell(row_num, 1).value
        if not date_val:
            continue

        analysis_date = str(date_val).split(' ')[0]

        # col 30(AD)부터 샘플 블록 (7컬럼 간격)
        for block_col in range(30, ws.max_column, 7):
            sample_name = ws.cell(row_num, block_col).value
            if not sample_name:
                break

            sample_ml = ws.cell(row_num, block_col + 1).value or ''
            absorbance = ws.cell(row_num, block_col + 2).value or ''
            dilution = ws.cell(row_num, block_col + 3).value or ''
            calibration_a = ws.cell(row_num, block_col + 4).value or ''
            result = ws.cell(row_num, block_col + 5).value or ''
            sn = ws.cell(row_num, block_col + 6).value or ''

            if not result or result == '0':
                continue

            # SN 기반 구분 판별
            division = '여수'
            if sn:
                sn_str = str(sn)
                if '율촌' in sn_str:
                    division = '율촌'
                    sn = sn_str.replace('[율촌]', '')
                elif '세풍' in sn_str:
                    division = '세풍'
                    sn = sn_str.replace('[세풍]', '')

            record = {
                '분석일': analysis_date,
                'SN': sn or '',
                '업체명': str(sample_name),
                '구분': division,
                '소스구분': '폐수배출업소',
                '시료명': str(sample_name),
                '분석자': '',
                '비고': '',
                '등록일시': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                '시료량': str(sample_ml),
                '흡광도': str(absorbance),
                '희석배수': str(dilution),
                '검량선_a': str(calibration_a),
                '결과': str(result)
            }
            records.append(record)

    print(f"✓ {len(records)}개 레코드 추출")

    if records:
        print(f"\nDB에 INSERT 중...")
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()

        inserted = 0
        for rec in records:
            try:
                query = """
                    INSERT IGNORE INTO `총질소_시험기록부`
                    (분석일, SN, 업체명, 구분, 소스구분, 시료명,
                     분석자, 비고, 등록일시,
                     시료량, 흡광도, 희석배수, 검량선_a, 결과)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                cursor.execute(query, (
                    rec['분석일'],
                    rec['SN'],
                    rec['업체명'],
                    rec['구분'],
                    rec['소스구분'],
                    rec['시료명'],
                    rec['분석자'],
                    rec['비고'],
                    rec['등록일시'],
                    rec['시료량'],
                    rec['흡광도'],
                    rec['희석배수'],
                    rec['검량선_a'],
                    rec['결과']
                ))
                inserted += 1

                if inserted % 100 == 0:
                    print(f"  {inserted}/{len(records)}...")
            except Exception as e:
                print(f"  ❌ Row 에러: {e}")

        conn.commit()
        print(f"✓ {inserted}개 행 INSERT 완료")
        cursor.close()
        conn.close()

except Exception as e:
    print(f"❌ 에러: {e}")
    import traceback
    traceback.print_exc()
