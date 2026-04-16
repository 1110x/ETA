#!/usr/bin/env python3
"""
처리시설 데이터 JSON → Excel 변환 스크립트
DB 테이블처럼 보이도록 포맷팅
"""

import json
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# siteCd → 시설명 매핑
FACILITY_MAP = {
    'PJT1020': '중흥',
    'PJT1021': '월내',
    'PJT1114': '4단계',
    'PJT1022': '율촌',
    'PJT1298': '세풍',
}

def extract_ocr_data(txt_file):
    """OCR 데이터만 추출"""
    with open(txt_file, 'r', encoding='utf-8') as f:
        content = f.read()

    # [OCR 데이터] 섹션만 추출
    ocr_pattern = r'\[OCR 데이터\][^\[]*?(\[\s*\{.*?\}\s*\])'
    matches = re.findall(ocr_pattern, content, re.DOTALL)

    all_records = []
    for match in matches:
        try:
            data = json.loads(match)
            all_records.extend(data if isinstance(data, list) else [data])
        except:
            pass

    return all_records

def get_facility_name(siteCd):
    """siteCd로 시설명 조회"""
    return FACILITY_MAP.get(siteCd, siteCd)

def convert_to_excel(txt_file, excel_file):
    """JSON 데이터를 Excel로 변환"""

    print(f"txt 파일 읽는 중: {txt_file}")
    records = extract_ocr_data(txt_file)
    print(f"추출된 레코드: {len(records)}개")

    # 시설별로 분류
    by_facility = {}
    for record in records:
        siteCd = record.get('siteCd', '')

        # 마스터 시료 스킵
        if siteCd in ['SITE053', 'SITE065', 'SITE066']:
            continue

        facility = get_facility_name(siteCd)
        if facility not in by_facility:
            by_facility[facility] = []
        by_facility[facility].append(record)

    # Excel 생성
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 삭제

    # 스타일 정의
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 시설별 시트 생성
    for facility, facility_records in by_facility.items():
        ws = wb.create_sheet(title=facility)

        # 헤더
        headers = ['날짜', '시료명', '분류', 'BOD', 'TOC', 'SS', 'T-N', 'T-P', '대장균군', '입력자', '입력일시']
        ws.append(headers)

        # 헤더 포맷
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # 데이터 행 추가
        for record in sorted(facility_records, key=lambda x: x.get('inputDate', '')):
            # BOD 값
            bod_d1 = record.get('bodD1')
            bod_d2 = record.get('bodD2')
            bod_value = ''
            if bod_d1 or bod_d2:
                if bod_d1 and bod_d2:
                    try:
                        bod_value = f"{(float(bod_d1) + float(bod_d2)) / 2:.2f}"
                    except:
                        bod_value = f"{bod_d1}/{bod_d2}"
                elif bod_d1:
                    bod_value = str(bod_d1)
                else:
                    bod_value = str(bod_d2)

            # TOC 값
            toc_ml = record.get('tocMl', '')
            toc_p = record.get('tocP', '')
            toc_value = f"{toc_ml}(×{toc_p})" if toc_ml else ''

            # SS 값
            ss_ml = record.get('ssMl', '')
            ss_p = record.get('ssP', '')
            ss_before = record.get('ssMgBefore', '')
            ss_after = record.get('ssMgAfter', '')
            ss_value = ''
            if ss_before and ss_after:
                ss_value = f"{ss_before}-{ss_after}"
            elif ss_ml:
                ss_value = f"{ss_ml}(×{ss_p})"

            # T-N 값
            tn_ml = record.get('tnMl', '')
            tn_p = record.get('tnP', '')
            tn_value = f"{tn_ml}(×{tn_p})" if tn_ml else ''

            # T-P 값
            tp_ml = record.get('tpMl', '')
            tp_p = record.get('tpP', '')
            tp_value = f"{tp_ml}(×{tp_p})" if tp_ml else ''

            # 대장균군 값
            coli_a = record.get('coliA', '')
            coli_b = record.get('coliB', '')
            coli_value = ''
            if coli_a or coli_b:
                coli_value = f"{coli_a}/{coli_b}" if coli_a and coli_b else (coli_a or coli_b)

            # 시료명
            sample_name = record.get('sampleCategoryNm', '')

            row = [
                record.get('inputDate', ''),
                sample_name,
                record.get('sampleCategory', ''),
                bod_value,
                toc_value,
                ss_value,
                tn_value,
                tp_value,
                coli_value,
                record.get('createUser', ''),
                record.get('createTime', '')
            ]

            ws.append(row)

        # 열 너비 자동 조정
        for col_num, header in enumerate(headers, 1):
            max_length = len(header)
            column = get_column_letter(col_num)

            for row_num in range(2, ws.max_row + 1):
                cell_value = str(ws.cell(row=row_num, column=col_num).value or '')
                max_length = max(max_length, len(cell_value))

            ws.column_dimensions[column].width = min(max_length + 2, 30)

        # 데이터 셀 포맷
        for row_num in range(2, ws.max_row + 1):
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 저장
    wb.save(excel_file)
    print(f"✓ Excel 저장 완료: {excel_file}")

if __name__ == '__main__':
    txt_file = Path.home() / 'Documents/ETA/Data/facility_data_last_week.txt'
    excel_file = Path.home() / 'Documents/ETA/Data/facility_data_last_week.xlsx'

    try:
        import openpyxl
    except:
        print("openpyxl 설치 중...")
        import subprocess
        subprocess.check_call(['pip3', 'install', 'openpyxl'])

    convert_to_excel(str(txt_file), str(excel_file))
