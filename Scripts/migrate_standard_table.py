"""
방류기준표 재구성 마이그레이션

1. 기존 DB `방류기준표` DROP
2. 엑셀 `방류기준표` 시트 → 새 `방류기준표` 테이블 생성
   - 컬럼: 구분(PK) + 방류기준 코드 (헤더 2행)
   - 행: 분석항목 67개
3. 새 테이블 `방류기준_매핑` 생성
4. 엑셀 `방류기준표 정리` → `방류기준_매핑` import
"""
import openpyxl
import pymysql

XLSM_PATH = "Data/★2026-센터청하★.xlsm"
DB = dict(host="1110s.synology.me", port=3306, db="eta_db",
          user="eta_user", password="1212xx!!AA", charset="utf8mb4")


def cell_val(cell):
    v = cell.value
    if v is None:
        return ""
    return str(v).strip()


def dedupe_columns(names):
    """중복 컬럼명에 _2, _3 접미사 자동 부여"""
    seen = {}
    result = []
    for n in names:
        if n in seen:
            seen[n] += 1
            result.append(f"{n}_{seen[n]}")
        else:
            seen[n] = 1
            result.append(n)
    return result


def main():
    print("=== 방류기준표 마이그레이션 시작 ===\n")
    conn = pymysql.connect(**DB)
    cur = conn.cursor()

    # ===== 1. 기존 방류기준표 DROP =====
    cur.execute("DROP TABLE IF EXISTS `방류기준표`")
    cur.execute("DROP TABLE IF EXISTS `방류기준_매핑`")
    conn.commit()
    print("[1] 기존 테이블 DROP 완료\n")

    # ===== 2. 엑셀 방류기준표 시트 읽기 =====
    wb = openpyxl.load_workbook(XLSM_PATH, data_only=True, read_only=True)
    ws = wb["방류기준표"]

    # 헤더 2행 → 방류기준 컬럼명 (col 1은 구분 컬럼, col 2부터 방류기준)
    raw_cols = []
    for c_idx in range(2, ws.max_column + 1):
        v = cell_val(ws.cell(2, c_idx))
        if v and v != "특례지역":
            # 특례지역은 건너뜀 (안내용)
            pass
        raw_cols.append((c_idx, v))

    # 유효 컬럼만 (빈값 제외)
    std_cols = [(c_idx, v) for c_idx, v in raw_cols if v]
    print(f"[2] 방류기준 컬럼 수: {len(std_cols)}개")

    # 컬럼명 dedupe
    col_names = [v for _, v in std_cols]
    col_names_deduped = dedupe_columns(col_names)
    print(f"    중복 처리 후: {len(col_names_deduped)}개")
    for i, (orig, new) in enumerate(zip(col_names, col_names_deduped)):
        if orig != new:
            print(f"      {orig} → {new}")

    # ===== 3. 방류기준표 CREATE =====
    col_defs = ["`_id` INT AUTO_INCREMENT PRIMARY KEY",
                "`구분` VARCHAR(100) NOT NULL"]
    for name in col_names_deduped:
        col_defs.append(f"`{name}` TEXT NULL DEFAULT NULL")

    create_sql = (f"CREATE TABLE `방류기준표` ({', '.join(col_defs)}) "
                  f"DEFAULT CHARSET=utf8mb4 ROW_FORMAT=DYNAMIC")
    cur.execute(create_sql)
    print(f"\n[3] 방류기준표 CREATE 완료 (컬럼 수: {len(col_names_deduped) + 2})")

    # ===== 4. 데이터 INSERT (3행부터 = 분석항목) =====
    inserted = 0
    for r_idx in range(3, ws.max_row + 1):
        구분 = cell_val(ws.cell(r_idx, 1))
        if not 구분:
            continue

        values = []
        for c_idx, _ in std_cols:
            v = cell_val(ws.cell(r_idx, c_idx))
            values.append(v if v else None)

        placeholders = ",".join(["%s"] * (1 + len(values)))
        col_list = "`구분`, " + ", ".join(f"`{n}`" for n in col_names_deduped)
        cur.execute(
            f"INSERT INTO `방류기준표` ({col_list}) VALUES ({placeholders})",
            tuple([구분] + values)
        )
        inserted += 1
    conn.commit()
    print(f"[4] 방류기준표 INSERT: {inserted}행\n")

    # ===== 5. 방류기준_매핑 CREATE =====
    cur.execute("""
        CREATE TABLE `방류기준_매핑` (
            `_id` INT AUTO_INCREMENT PRIMARY KEY,
            `업체명` VARCHAR(100) NOT NULL,
            `시료명` VARCHAR(200) NOT NULL,
            `방류기준코드` VARCHAR(100) NULL,
            INDEX idx_시료명 (`시료명`)
        ) DEFAULT CHARSET=utf8mb4
    """)
    print("[5] 방류기준_매핑 CREATE 완료")

    # ===== 6. 방류기준표 정리 → 방류기준_매핑 import =====
    ws2 = wb["방류기준표 정리"]
    # 헤더 1행: [업체명 | "방류기준"] 쌍 반복
    company_cols = []  # (업체명, 시료명 컬럼번호, 방류기준 컬럼번호)
    for c in range(1, ws2.max_column + 1, 2):
        업체명 = cell_val(ws2.cell(1, c))
        if 업체명 and 업체명 != "방류기준":
            company_cols.append((업체명, c, c + 1))

    print(f"    업체 수: {len(company_cols)}")

    mapping_inserted = 0
    for 업체명, sample_col, std_col in company_cols:
        for r_idx in range(2, ws2.max_row + 1):
            시료명 = cell_val(ws2.cell(r_idx, sample_col))
            방류기준코드 = cell_val(ws2.cell(r_idx, std_col))
            if not 시료명:
                continue
            cur.execute(
                "INSERT INTO `방류기준_매핑` (`업체명`, `시료명`, `방류기준코드`) VALUES (%s, %s, %s)",
                (업체명, 시료명, 방류기준코드 if 방류기준코드 else None)
            )
            mapping_inserted += 1
    conn.commit()
    print(f"[6] 방류기준_매핑 INSERT: {mapping_inserted}행\n")

    wb.close()
    conn.close()
    print("=== 마이그레이션 완료 ===")


if __name__ == "__main__":
    main()
