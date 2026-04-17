"""
GS CODE 시트 → GS_CODE 테이블 import

엑셀 구조:
- 행2: 헤더 (코드, 오염물질명, 단위, [빈], [빈], W1, W2, ..., W5, #W6 ~ #W24)
- 행3~94: 데이터 (92개 항목)
- col 4, 5는 비어있음 (엑셀 시트 디자인상)
"""
import openpyxl
import pymysql

XLSM = "Data/★2026-센터청하★.xlsm"
DB = dict(host="1110s.synology.me", port=3306, db="eta_db",
          user="eta_user", password="1212xx!!AA", charset="utf8mb4")

# 배출구 컬럼들 (col 6 ~ col 29 → 24개)
WELL_COLS = ["W1", "W2", "W3", "W4", "W5"] + [f"#W{i}" for i in range(6, 25)]


def main():
    print("=== GS_CODE 마이그레이션 시작 ===\n")
    conn = pymysql.connect(**DB)
    cur = conn.cursor()

    # 기존 테이블 DROP
    cur.execute("DROP TABLE IF EXISTS `GS_CODE`")

    # CREATE
    col_defs = [
        "`_id` INT AUTO_INCREMENT PRIMARY KEY",
        "`코드` INT NOT NULL",
        "`오염물질명` VARCHAR(200) NULL",
        "`단위` VARCHAR(20) NULL",
    ]
    for w in WELL_COLS:
        col_defs.append(f"`{w}` TEXT NULL DEFAULT NULL")
    create_sql = (f"CREATE TABLE `GS_CODE` ({', '.join(col_defs)}) "
                  f"DEFAULT CHARSET=utf8mb4 ROW_FORMAT=DYNAMIC")
    cur.execute(create_sql)
    print(f"[1] GS_CODE 테이블 CREATE (컬럼 {len(col_defs)}개)")

    # 엑셀 → INSERT
    wb = openpyxl.load_workbook(XLSM, data_only=True, read_only=True)
    ws = wb["GS CODE"]

    inserted = 0
    for r in range(3, 95):
        code = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        unit = ws.cell(r, 3).value
        if code is None:
            continue
        try:
            code_int = int(code)
        except Exception:
            continue

        # 배출구별 값 (col 6~29)
        well_vals = []
        for i, w in enumerate(WELL_COLS):
            v = ws.cell(r, 6 + i).value
            well_vals.append(None if v is None else str(v).strip())

        col_list = "`코드`, `오염물질명`, `단위`, " + ", ".join(f"`{w}`" for w in WELL_COLS)
        placeholders = ",".join(["%s"] * (3 + len(WELL_COLS)))
        cur.execute(
            f"INSERT INTO `GS_CODE` ({col_list}) VALUES ({placeholders})",
            tuple([code_int, name, unit] + well_vals)
        )
        inserted += 1

    conn.commit()
    print(f"[2] INSERT 완료: {inserted}행")
    wb.close()
    conn.close()
    print("\n=== 완료 ===")


if __name__ == "__main__":
    main()
