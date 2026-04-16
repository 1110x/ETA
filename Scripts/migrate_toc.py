"""
TOC(TCIC)-DATA -> 총_유기탄소_TCIC_시험기록부
TOC(NPOC)-DATA -> 총_유기탄소_NPOC_시험기록부

TCIC 고정영역:
col3~7:   TC ST-01~05 mgL  -> ST01~05_mgL_TCIC
col8~12:  TC ST-01~05 abs  -> ST01~05_abs_TCIC
col13~15: TC 기울기/절편/R2 -> 기울기_TCIC/절편_TCIC/R2_TCIC
col16~20: IC ST-01~05 mgL  -> ST01~05_mgL_IC
col21~25: IC ST-01~05 abs  -> ST01~05_abs_IC
col26~28: IC 기울기/절편/R2 -> 기울기_IC/절편_IC/R2_IC

TCIC 샘플 블록 (col30부터 8컬럼 간격):
+0=시료명, +1=시료량(미저장), +2=TCcon, +3=ICcon,
+4=농도_TCIC, +5=희석배수_TCIC, +6=결과_TCIC, +7=SN

NPOC 고정영역:
col3~7:   ST-01~05 mgL  -> ST01~05_mgL_NPOC
col8~12:  ST-01~05 abs  -> ST01~05_abs_NPOC
col13~15: 기울기/절편/R2 -> 기울기_NPOC/절편_NPOC/R2_NPOC

NPOC 샘플 블록 (col30부터 8컬럼 간격):
+0=시료명, +1=시료량_NPOC, +2=흡광도_NPOC, +3=희석배수_NPOC,
+4=검량선_a_NPOC, +5=농도_NPOC, +6=결과_NPOC, +7=SN
"""
import openpyxl
import pymysql
from datetime import datetime, date, timedelta

XLSM_PATH = "Docs/CHUNGHA-김가린.xlsm"
DB = dict(host="1110s.synology.me", port=3306, db="eta_db",
          user="eta_user", password="1212xx!!AA", charset="utf8mb4")

TCIC_TBL = "총_유기탄소_TCIC_시험기록부"
NPOC_TBL = "총_유기탄소_NPOC_시험기록부"


def get_val(cell):
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    return "" if s == "-" else s


def try_parse_date(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    try:
        days = int(float(s))
        if days > 59:
            days -= 1
        return (datetime(1899, 12, 30) + timedelta(days=days)).strftime("%Y-%m-%d")
    except Exception:
        pass
    if len(s) == 10 and s[4] == '-':
        return s
    try:
        s2 = s.split(". ")[0].strip().rstrip(".")
        return datetime.strptime(s2, "%Y. %m. %d").strftime("%Y-%m-%d")
    except Exception:
        pass
    return None


def normalize_sn(sn):
    parts = sn.split("-")
    if len(parts) != 3:
        return sn
    try:
        return "%02d-%02d-%02d" % (int(parts[0]), int(parts[1]), int(parts[2]))
    except Exception:
        return sn


def parse_sn(raw):
    s = raw.strip()
    if not s:
        return None, None
    gub = "여수"
    if s.startswith("[율촌]"):
        gub = "율촌"
        s = s[5:]
    elif s.startswith("[세풍]"):
        gub = "세풍"
        s = s[5:]
    try:
        float(s)
        return None, None
    except Exception:
        pass
    return normalize_sn(s), gub


def get_date_col(conn):
    with conn.cursor() as cur:
        cur.execute("SHOW COLUMNS FROM `" + TCIC_TBL + "`")
        cols = [row[0] for row in cur.fetchall()]
    return next(c for c in cols if '일' in c and '등록' not in c)


def upsert_tcic(conn, col_date, 날짜, sn, 업체명, 구분,
                tccon, iccon, 농도, 희석, 결과,
                tc_mgl, tc_abs, tc_curve,
                ic_mgl, ic_abs, ic_curve):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sel = ("SELECT id FROM `" + TCIC_TBL + "` "
           "WHERE LEFT(`" + col_date + "`,10)=%s AND SN=%s LIMIT 1")
    with conn.cursor() as cur:
        cur.execute(sel, (날짜, sn))
        row = cur.fetchone()
        if row:
            upd = ("UPDATE `" + TCIC_TBL + "` SET "
                   "TCcon=%s, ICcon=%s, 농도_TCIC=%s, 희석배수_TCIC=%s, 결과_TCIC=%s, "
                   "ST01_mgL_TCIC=%s, ST02_mgL_TCIC=%s, ST03_mgL_TCIC=%s, ST04_mgL_TCIC=%s, ST05_mgL_TCIC=%s, "
                   "ST01_abs_TCIC=%s, ST02_abs_TCIC=%s, ST03_abs_TCIC=%s, ST04_abs_TCIC=%s, ST05_abs_TCIC=%s, "
                   "기울기_TCIC=%s, 절편_TCIC=%s, R2_TCIC=%s, "
                   "ST01_mgL_IC=%s, ST02_mgL_IC=%s, ST03_mgL_IC=%s, ST04_mgL_IC=%s, ST05_mgL_IC=%s, "
                   "ST01_abs_IC=%s, ST02_abs_IC=%s, ST03_abs_IC=%s, ST04_abs_IC=%s, ST05_abs_IC=%s, "
                   "기울기_IC=%s, 절편_IC=%s, R2_IC=%s, "
                   "등록일시=%s WHERE id=%s")
            cur.execute(upd, (
                tccon, iccon, 농도, 희석, 결과,
                tc_mgl[0], tc_mgl[1], tc_mgl[2], tc_mgl[3], tc_mgl[4],
                tc_abs[0], tc_abs[1], tc_abs[2], tc_abs[3], tc_abs[4],
                tc_curve[0], tc_curve[1], tc_curve[2],
                ic_mgl[0], ic_mgl[1], ic_mgl[2], ic_mgl[3], ic_mgl[4],
                ic_abs[0], ic_abs[1], ic_abs[2], ic_abs[3], ic_abs[4],
                ic_curve[0], ic_curve[1], ic_curve[2],
                now, row[0]
            ))
        else:
            ins = ("INSERT INTO `" + TCIC_TBL + "` "
                   "(`" + col_date + "`, SN, 업체명, 시료명, 구분, 소스구분, "
                   "TCcon, ICcon, 농도_TCIC, 희석배수_TCIC, 결과_TCIC, "
                   "ST01_mgL_TCIC, ST02_mgL_TCIC, ST03_mgL_TCIC, ST04_mgL_TCIC, ST05_mgL_TCIC, "
                   "ST01_abs_TCIC, ST02_abs_TCIC, ST03_abs_TCIC, ST04_abs_TCIC, ST05_abs_TCIC, "
                   "기울기_TCIC, 절편_TCIC, R2_TCIC, "
                   "ST01_mgL_IC, ST02_mgL_IC, ST03_mgL_IC, ST04_mgL_IC, ST05_mgL_IC, "
                   "ST01_abs_IC, ST02_abs_IC, ST03_abs_IC, ST04_abs_IC, ST05_abs_IC, "
                   "기울기_IC, 절편_IC, R2_IC, 등록일시) "
                   "VALUES (%s,%s,%s,%s,%s,%s,"
                   "%s,%s,%s,%s,%s,"
                   "%s,%s,%s,%s,%s,"
                   "%s,%s,%s,%s,%s,"
                   "%s,%s,%s,"
                   "%s,%s,%s,%s,%s,"
                   "%s,%s,%s,%s,%s,"
                   "%s,%s,%s,%s)")
            cur.execute(ins, (
                날짜, sn, 업체명, 업체명, 구분, "폐수배출업소",
                tccon, iccon, 농도, 희석, 결과,
                tc_mgl[0], tc_mgl[1], tc_mgl[2], tc_mgl[3], tc_mgl[4],
                tc_abs[0], tc_abs[1], tc_abs[2], tc_abs[3], tc_abs[4],
                tc_curve[0], tc_curve[1], tc_curve[2],
                ic_mgl[0], ic_mgl[1], ic_mgl[2], ic_mgl[3], ic_mgl[4],
                ic_abs[0], ic_abs[1], ic_abs[2], ic_abs[3], ic_abs[4],
                ic_curve[0], ic_curve[1], ic_curve[2],
                now
            ))


def upsert_npoc(conn, col_date, 날짜, sn, 업체명, 구분,
                시료량, 흡광도, 희석, 검량선, 농도, 결과,
                st_mgl, st_abs, curve):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sel = ("SELECT id FROM `" + NPOC_TBL + "` "
           "WHERE LEFT(`" + col_date + "`,10)=%s AND SN=%s LIMIT 1")
    with conn.cursor() as cur:
        cur.execute(sel, (날짜, sn))
        row = cur.fetchone()
        if row:
            upd = ("UPDATE `" + NPOC_TBL + "` SET "
                   "시료량_NPOC=%s, 흡광도_NPOC=%s, 희석배수_NPOC=%s, "
                   "검량선_a_NPOC=%s, 농도_NPOC=%s, 결과_NPOC=%s, "
                   "기울기_NPOC=%s, 절편_NPOC=%s, R2_NPOC=%s, "
                   "ST01_mgL_NPOC=%s, ST02_mgL_NPOC=%s, ST03_mgL_NPOC=%s, ST04_mgL_NPOC=%s, ST05_mgL_NPOC=%s, "
                   "ST01_abs_NPOC=%s, ST02_abs_NPOC=%s, ST03_abs_NPOC=%s, ST04_abs_NPOC=%s, ST05_abs_NPOC=%s, "
                   "등록일시=%s WHERE id=%s")
            cur.execute(upd, (
                시료량, 흡광도, 희석, 검량선, 농도, 결과,
                curve[0], curve[1], curve[2],
                st_mgl[0], st_mgl[1], st_mgl[2], st_mgl[3], st_mgl[4],
                st_abs[0], st_abs[1], st_abs[2], st_abs[3], st_abs[4],
                now, row[0]
            ))
        else:
            ins = ("INSERT INTO `" + NPOC_TBL + "` "
                   "(`" + col_date + "`, SN, 업체명, 시료명, 구분, 소스구분, "
                   "시료량_NPOC, 흡광도_NPOC, 희석배수_NPOC, 검량선_a_NPOC, 농도_NPOC, 결과_NPOC, "
                   "기울기_NPOC, 절편_NPOC, R2_NPOC, "
                   "ST01_mgL_NPOC, ST02_mgL_NPOC, ST03_mgL_NPOC, ST04_mgL_NPOC, ST05_mgL_NPOC, "
                   "ST01_abs_NPOC, ST02_abs_NPOC, ST03_abs_NPOC, ST04_abs_NPOC, ST05_abs_NPOC, "
                   "등록일시) "
                   "VALUES (%s,%s,%s,%s,%s,%s,"
                   "%s,%s,%s,%s,%s,%s,"
                   "%s,%s,%s,"
                   "%s,%s,%s,%s,%s,"
                   "%s,%s,%s,%s,%s,"
                   "%s)")
            cur.execute(ins, (
                날짜, sn, 업체명, 업체명, 구분, "폐수배출업소",
                시료량, 흡광도, 희석, 검량선, 농도, 결과,
                curve[0], curve[1], curve[2],
                st_mgl[0], st_mgl[1], st_mgl[2], st_mgl[3], st_mgl[4],
                st_abs[0], st_abs[1], st_abs[2], st_abs[3], st_abs[4],
                now
            ))


def load_tcic(wb, conn, col_date):
    ws = wb["TOC(TCIC)-DATA"]
    loaded = 0
    skipped = 0

    for row_num, r in enumerate(ws.iter_rows(min_row=2), start=2):
        날짜 = try_parse_date(r[0].value)
        if not 날짜:
            continue

        tc_mgl   = [get_val(r[i]) for i in range(2, 7)]
        tc_abs   = [get_val(r[i]) for i in range(7, 12)]
        tc_curve = [get_val(r[12]), get_val(r[13]), get_val(r[14])]
        ic_mgl   = [get_val(r[i]) for i in range(15, 20)]
        ic_abs   = [get_val(r[i]) for i in range(20, 25)]
        ic_curve = [get_val(r[25]), get_val(r[26]), get_val(r[27])]

        for block in range(80):
            idx = 29 + block * 8
            if idx >= len(r):
                break
            시료명 = get_val(r[idx])
            if not 시료명:
                break
            결과   = get_val(r[idx + 6]) if idx + 6 < len(r) else ""
            sn_raw = get_val(r[idx + 7]) if idx + 7 < len(r) else ""
            if not 결과 or 결과 == "0":
                skipped += 1
                continue
            sn, 구분 = parse_sn(sn_raw)
            if sn is None:
                skipped += 1
                continue
            tccon = get_val(r[idx + 2]) if idx + 2 < len(r) else ""
            iccon = get_val(r[idx + 3]) if idx + 3 < len(r) else ""
            농도  = get_val(r[idx + 4]) if idx + 4 < len(r) else ""
            희석  = get_val(r[idx + 5]) if idx + 5 < len(r) else ""
            try:
                upsert_tcic(conn, col_date, 날짜, sn, 시료명, 구분,
                            tccon, iccon, 농도, 희석, 결과,
                            tc_mgl, tc_abs, tc_curve,
                            ic_mgl, ic_abs, ic_curve)
                loaded += 1
                if loaded % 200 == 0:
                    conn.commit()
                    print("    [TCIC] %d건 처리 중... (행%d)" % (loaded, row_num))
            except Exception as e:
                print("    오류 행%d 블록%d: %s" % (row_num, block, e))

    conn.commit()
    print("  [TCIC] %d건 저장, %d건 스킵" % (loaded, skipped))
    return loaded


def load_npoc(wb, conn, col_date):
    ws = wb["TOC(NPOC)-DATA"]
    loaded = 0
    skipped = 0

    for row_num, r in enumerate(ws.iter_rows(min_row=2), start=2):
        날짜 = try_parse_date(r[0].value)
        if not 날짜:
            continue

        st_mgl = [get_val(r[i]) for i in range(2, 7)]
        st_abs = [get_val(r[i]) for i in range(7, 12)]
        curve  = [get_val(r[12]), get_val(r[13]), get_val(r[14])]

        for block in range(80):
            idx = 29 + block * 8
            if idx >= len(r):
                break
            시료명 = get_val(r[idx])
            if not 시료명:
                break
            결과   = get_val(r[idx + 6]) if idx + 6 < len(r) else ""
            sn_raw = get_val(r[idx + 7]) if idx + 7 < len(r) else ""
            if not 결과 or 결과 == "0":
                skipped += 1
                continue
            sn, 구분 = parse_sn(sn_raw)
            if sn is None:
                skipped += 1
                continue
            시료량 = get_val(r[idx + 1]) if idx + 1 < len(r) else ""
            흡광도 = get_val(r[idx + 2]) if idx + 2 < len(r) else ""
            희석   = get_val(r[idx + 3]) if idx + 3 < len(r) else ""
            검량선 = get_val(r[idx + 4]) if idx + 4 < len(r) else ""
            농도   = get_val(r[idx + 5]) if idx + 5 < len(r) else ""
            try:
                upsert_npoc(conn, col_date, 날짜, sn, 시료명, 구분,
                            시료량, 흡광도, 희석, 검량선, 농도, 결과,
                            st_mgl, st_abs, curve)
                loaded += 1
                if loaded % 200 == 0:
                    conn.commit()
                    print("    [NPOC] %d건 처리 중... (행%d)" % (loaded, row_num))
            except Exception as e:
                print("    오류 행%d 블록%d: %s" % (row_num, block, e))

    conn.commit()
    print("  [NPOC] %d건 저장, %d건 스킵" % (loaded, skipped))
    return loaded


def main():
    print("TOC 마이그레이션 시작")
    conn = pymysql.connect(**DB)
    col_date = get_date_col(conn)
    print("날짜 컬럼: %s" % repr(col_date))

    with conn.cursor() as cur:
        cur.execute("DELETE FROM `" + TCIC_TBL + "`")
        cur.execute("DELETE FROM `" + NPOC_TBL + "`")
    conn.commit()
    print("기존 데이터 삭제 완료")

    wb = openpyxl.load_workbook(XLSM_PATH, data_only=True, read_only=True)
    try:
        total = 0
        total += load_tcic(wb, conn, col_date)
        total += load_npoc(wb, conn, col_date)
        print("\n완료: 총 %d건" % total)
    finally:
        conn.close()
        wb.close()


if __name__ == "__main__":
    main()
