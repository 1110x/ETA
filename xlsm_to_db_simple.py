#!/usr/bin/env python3
"""
청하 XLSM BOD-DATA → 생물학적_산소요구량_시험기록부
정확한 컬럼 매핑으로 데이터 추출 및 INSERT
"""

import openpyxl
from datetime import datetime
import mysql.connector

DB_CONFIG = {
    'host': '1110s.synology.me',
    'user': 'eta_user',
    'password': '1212xx!!AA',
    'database': 'eta_db',
    'charset': 'utf8mb4'
}

xlsm_file = '/Users/azrael/Documents/ETA/Docs/CHUNGHA-김가린.xlsm'

try:
    print("XLSM 파일 로드 중...")
    wb = openpyxl.load_workbook(xlsm_file)
    ws = wb['BOD-DATA']
    print(f"✓ BOD-DATA 시트 로드 (max_row: {ws.max_row})")

    records = []

    # 모든 행 처리 (2~2804)
    for row_num in range(2, ws.max_row + 1):
        date_val = ws.cell(row_num, 1).value
        if not date_val:
            continue

        analysis_date = str(date_val).split(' ')[0]

        # 고정 영역 (B~K 열)
        analyzer1 = ws.cell(row_num, 2).value or ''
        analyzer2 = ws.cell(row_num, 3).value or ''
        seed_sample_ml = ws.cell(row_num, 4).value or ''
        do_15min = ws.cell(row_num, 5).value or ''
        do_5day = ws.cell(row_num, 6).value or ''
        seed_bod = ws.cell(row_num, 7).value or ''
        seed_content = ws.cell(row_num, 8).value or ''
        dilution_sample_ml = ws.cell(row_num, 9).value or ''
        seed_d1 = ws.cell(row_num, 10).value or ''
        seed_d2 = ws.cell(row_num, 11).value or ''

        # col 30(AD)부터 샘플 블록
        for block_col in range(30, ws.max_column, 8):
            sample_name = ws.cell(row_num, block_col).value
            if not sample_name:
                break

            sample_ml = ws.cell(row_num, block_col + 1).value or ''
            d1 = ws.cell(row_num, block_col + 2).value or ''
            d2 = ws.cell(row_num, block_col + 3).value or ''
            dilution = ws.cell(row_num, block_col + 5).value or ''
            result = ws.cell(row_num, block_col + 6).value or ''
            sn = ws.cell(row_num, block_col + 7).value or ''

            if not result or result == '0':
                continue

            # SN 기반 구분 판별
            division = '여수'
            if sn:
                sn_str = str(sn)
                if '율촌' in sn_str:
                    division = '율촌'
                    sn = sn_str.replace('[율촌]', '')
                elif '세풍' in sn_str:
                    division = '세풍'
                    sn = sn_str.replace('[세풍]', '')

            record = {
                '분석일': analysis_date,
                'SN': sn or '',
                '업체명': str(sample_name),
                '구분': division,
                '소스구분': '폐수배출업소',
                '시료명': str(sample_name),
                '분석자1': str(analyzer1),
                '분석자2': str(analyzer2),
                '식종시료량': str(seed_sample_ml),
                '15min_DO': str(do_15min),
                '5Day_DO': str(do_5day),
                '식종BOD': str(seed_bod),
                '식종함유량': str(seed_content),
                '희석수시료량': str(dilution_sample_ml),
                '식종D1': str(seed_d1),
                '식종D2': str(seed_d2),
                '시료량': str(sample_ml),
                'D1': str(d1),
                'D2': str(d2),
                '희석배수': str(dilution),
                '결과': str(result),
                '의뢰명': '',
                '시료분석': '',
                '비고': '',
                '등록일시': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            records.append(record)

    print(f"✓ {len(records)}개 레코드 추출")

    if records:
        print("\n[ 샘플 데이터 ]")
        for i, rec in enumerate(records[:5]):
            print(f"  {i+1}. {rec['업체명']} - {rec['결과']}")

        print(f"\nDB에 INSERT 중...")
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()

        inserted = 0
        for rec in records:
            try:
                query = """
                    INSERT IGNORE INTO `생물화학적_산소요구량_시험기록부`
                    (분석일, SN, 업체명, 구분, 소스구분, 시료명,
                     분석자1, 분석자2, 식종시료량, `15min_DO`, `5Day_DO`,
                     식종BOD, 식종함유량, 희석수시료량, 식종D1, 식종D2,
                     시료량, D1, D2, 희석배수, 결과,
                     의뢰명, 시료분석, 비고, 등록일시)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                cursor.execute(query, (
                    rec['분석일'],
                    rec['SN'],
                    rec['업체명'],
                    rec['구분'],
                    rec['소스구분'],
                    rec['시료명'],
                    rec['분석자1'],
                    rec['분석자2'],
                    rec['식종시료량'],
                    rec['15min_DO'],
                    rec['5Day_DO'],
                    rec['식종BOD'],
                    rec['식종함유량'],
                    rec['희석수시료량'],
                    rec['식종D1'],
                    rec['식종D2'],
                    rec['시료량'],
                    rec['D1'],
                    rec['D2'],
                    rec['희석배수'],
                    rec['결과'],
                    rec['의뢰명'],
                    rec['시료분석'],
                    rec['비고'],
                    rec['등록일시']
                ))
                inserted += 1

                # 진행 표시
                if inserted % 100 == 0:
                    print(f"  {inserted}/{len(records)}...")
            except Exception as e:
                print(f"  ❌ Row 에러: {e}")

        conn.commit()
        print(f"✓ {inserted}개 행 INSERT 완료")
        cursor.close()
        conn.close()

except Exception as e:
    print(f"❌ 에러: {e}")
    import traceback
    traceback.print_exc()
