#!/usr/bin/env python3
"""
청하 XLSM TOC-DATA (TCIC + NPOC) → 총_유기탄소_시험기록부
"""

import openpyxl
from datetime import datetime
import mysql.connector

DB_CONFIG = {
    'host': '1110s.synology.me',
    'user': 'eta_user',
    'password': '1212xx!!AA',
    'database': 'eta_db',
    'charset': 'utf8mb4'
}

xlsm_file = '/Users/azrael/Documents/ETA/Docs/CHUNGHA-김가린.xlsm'

try:
    print("XLSM 파일 로드 중...")
    wb = openpyxl.load_workbook(xlsm_file)

    records = []

    # TCIC 먼저 처리
    ws_tcic = wb['TOC(TCIC)-DATA']
    ws_npoc = wb['TOC(NPOC)-DATA']

    print(f"✓ TOC(TCIC)-DATA 시트 로드 (max_row: {ws_tcic.max_row})")
    print(f"✓ TOC(NPOC)-DATA 시트 로드 (max_row: {ws_npoc.max_row})")

    # TCIC 데이터 처리
    for row_num in range(2, ws_tcic.max_row + 1):
        date_val = ws_tcic.cell(row_num, 1).value
        if not date_val:
            continue

        analysis_date = str(date_val).split(' ')[0]

        for block_col in range(30, ws_tcic.max_column, 8):
            sample_name = ws_tcic.cell(row_num, block_col).value
            if not sample_name:
                break

            sample_ml = ws_tcic.cell(row_num, block_col + 1).value or ''
            absorbance = ws_tcic.cell(row_num, block_col + 2).value or ''
            dilution = ws_tcic.cell(row_num, block_col + 3).value or ''
            calibration_a = ws_tcic.cell(row_num, block_col + 4).value or ''
            concentration = ws_tcic.cell(row_num, block_col + 5).value or ''
            result = ws_tcic.cell(row_num, block_col + 6).value or ''
            sn = ws_tcic.cell(row_num, block_col + 7).value or ''

            if not result or result == '0':
                continue

            division = '여수'
            if sn:
                sn_str = str(sn)
                if '율촌' in sn_str:
                    division = '율촌'
                    sn = sn_str.replace('[율촌]', '')
                elif '세풍' in sn_str:
                    division = '세풍'
                    sn = sn_str.replace('[세풍]', '')

            record = {
                '분석일': analysis_date,
                'SN': sn or '',
                '업체명': str(sample_name),
                '구분': division,
                '소스구분': '폐수배출업소',
                '시료명': str(sample_name),
                '분석자': '',
                '비고': '',
                '등록일시': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                '시료량': str(sample_ml),
                '흡광도': str(absorbance),
                '희석배수': str(dilution),
                '검량선_a': str(calibration_a),
                '농도': str(concentration),
                '결과': str(result),
                '방법': 'TCIC'
            }
            records.append(record)

    # NPOC 데이터도 추가 (결과가 있는 것들)
    for row_num in range(2, ws_npoc.max_row + 1):
        date_val = ws_npoc.cell(row_num, 1).value
        if not date_val:
            continue

        analysis_date = str(date_val).split(' ')[0]

        for block_col in range(30, ws_npoc.max_column, 8):
            sample_name = ws_npoc.cell(row_num, block_col).value
            if not sample_name:
                break

            result = ws_npoc.cell(row_num, block_col + 6).value or ''
            if not result or result == '0':
                continue

            sample_ml = ws_npoc.cell(row_num, block_col + 1).value or ''
            absorbance = ws_npoc.cell(row_num, block_col + 2).value or ''
            dilution = ws_npoc.cell(row_num, block_col + 3).value or ''
            calibration_a = ws_npoc.cell(row_num, block_col + 4).value or ''
            concentration = ws_npoc.cell(row_num, block_col + 5).value or ''
            sn = ws_npoc.cell(row_num, block_col + 7).value or ''

            division = '여수'
            if sn:
                sn_str = str(sn)
                if '율촌' in sn_str:
                    division = '율촌'
                    sn = sn_str.replace('[율촌]', '')
                elif '세풍' in sn_str:
                    division = '세풍'
                    sn = sn_str.replace('[세풍]', '')

            record = {
                '분석일': analysis_date,
                'SN': sn or '',
                '업체명': str(sample_name),
                '구분': division,
                '소스구분': '폐수배출업소',
                '시료명': str(sample_name),
                '분석자': '',
                '비고': '',
                '등록일시': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                '시료량': str(sample_ml),
                '흡광도': str(absorbance),
                '희석배수': str(dilution),
                '검량선_a': str(calibration_a),
                '농도': str(concentration),
                '결과': str(result),
                '방법': 'NPOC'
            }
            records.append(record)

    print(f"✓ {len(records)}개 레코드 추출")

    if records:
        print(f"\nDB에 INSERT 중...")
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()

        inserted = 0
        for rec in records:
            try:
                query = """
                    INSERT IGNORE INTO `총_유기탄소_시험기록부`
                    (분석일, SN, 업체명, 구분, 소스구분, 시료명,
                     분석자, 비고, 등록일시,
                     시료량, 흡광도, 희석배수, 검량선_a, 농도, 결과, 방법)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                cursor.execute(query, (
                    rec['분석일'], rec['SN'], rec['업체명'], rec['구분'],
                    rec['소스구분'], rec['시료명'], rec['분석자'], rec['비고'],
                    rec['등록일시'], rec['시료량'], rec['흡광도'], rec['희석배수'],
                    rec['검량선_a'], rec['농도'], rec['결과'], rec['방법']
                ))
                inserted += 1

                if inserted % 100 == 0:
                    print(f"  {inserted}/{len(records)}...")
            except Exception as e:
                pass

        conn.commit()
        print(f"✓ {inserted}개 행 INSERT 완료")
        cursor.close()
        conn.close()

except Exception as e:
    print(f"❌ 에러: {e}")
