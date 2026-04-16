#!/usr/bin/env python3
"""
처리시설 데이터 JSON → Excel 변환 스크립트 v2
각 분석항목별로 올바른 결과값 추출
"""

import json
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# siteCd → 시설명 매핑
FACILITY_MAP = {
    'PJT1020': '중흥',
    'PJT1021': '월내',
    'PJT1114': '4단계',
    'PJT1022': '율촌',
    'PJT1298': '세풍',
}

def extract_data_by_section(txt_file):
    """TXT 파일에서 섹션별로 데이터 추출"""
    with open(txt_file, 'r', encoding='utf-8') as f:
        content = f.read()

    sections = {}

    # 각 섹션별 추출
    for section_name in ['OCR', 'SS', 'T-N', 'T-P', 'TOC', 'Coli']:
        if section_name == 'OCR':
            pattern = r'\[OCR 데이터\][^\[]*?(\[\s*\{.*?\}\s*\])'
        elif section_name == 'SS':
            pattern = r'\[SS 데이터\][^\[]*?(\[\s*\{.*?\}\s*\])'
        elif section_name == 'T-N':
            pattern = r'\[T-N 데이터\][^\[]*?(\[\s*\{.*?\}\s*\])'
        elif section_name == 'T-P':
            pattern = r'\[T-P 데이터\][^\[]*?(\[\s*\{.*?\}\s*\])'
        elif section_name == 'TOC':
            pattern = r'\[TOC 데이터\][^\[]*?(\[\s*\{.*?\}\s*\])'
        elif section_name == 'Coli':
            pattern = r'\[대장균군 데이터\][^\[]*?(\[\s*\{.*?\}\s*\])'

        matches = re.findall(pattern, content, re.DOTALL)
        if matches:
            try:
                data = json.loads(matches[0])
                sections[section_name] = data if isinstance(data, list) else [data]
            except:
                sections[section_name] = []

    return sections

def get_facility_name(siteCd):
    """siteCd로 시설명 조회"""
    return FACILITY_MAP.get(siteCd, siteCd)

def build_result_map(sections):
    """섹션 데이터를 (inputDate, siteCd, sampleCategory) → 결과값 맵으로 구성"""
    result_map = {}

    # BOD, TOC, T-N, T-P: OCR 데이터에서 추출
    if 'OCR' in sections:
        for item in sections['OCR']:
            key = (item.get('inputDate'), item.get('siteCd'), item.get('sampleCategory'))
            if key not in result_map:
                result_map[key] = {}

            # BOD: D1, D2 평균 (bodNo가 있으면 유효한 데이터)
            if item.get('bodNo'):
                bod_d1 = item.get('bodD1')
                bod_d2 = item.get('bodD2')
                if bod_d1 and bod_d2:
                    try:
                        result_map[key]['BOD'] = f"{(float(bod_d1) + float(bod_d2)) / 2:.2f}"
                    except:
                        result_map[key]['BOD'] = ''

            # TOC: Ml(×P) 형식
            toc_ml = item.get('tocMl')
            toc_p = item.get('tocP')
            if toc_ml:
                result_map[key]['TOC'] = f"{toc_ml}(×{toc_p})" if toc_p else toc_ml

    # SS: ssMgBefore, ssMgAfter 무게 차이
    if 'SS' in sections:
        for item in sections['SS']:
            key = (item.get('inputDate'), item.get('siteCd'), item.get('sampleCategory'))
            if key not in result_map:
                result_map[key] = {}

            ss_before = item.get('ssMgBefore')
            ss_after = item.get('ssMgAfter')
            if ss_before and ss_after:
                try:
                    result_map[key]['SS'] = f"{ss_before}-{ss_after}"
                except:
                    pass

    # T-N: tnDensityMg (결과값)
    if 'T-N' in sections:
        for item in sections['T-N']:
            key = (item.get('inputDate'), item.get('siteCd'), item.get('sampleCategory'))
            if key not in result_map:
                result_map[key] = {}

            tn_density = item.get('tnDensityMg')
            if tn_density:
                result_map[key]['T-N'] = str(tn_density)

    # T-P: tpDensityMg (결과값)
    if 'T-P' in sections:
        for item in sections['T-P']:
            key = (item.get('inputDate'), item.get('siteCd'), item.get('sampleCategory'))
            if key not in result_map:
                result_map[key] = {}

            tp_density = item.get('tpDensityMg')
            if tp_density:
                result_map[key]['T-P'] = str(tp_density)

    # TOC: tocTcMg 또는 tocIcMg (결과값)
    if 'TOC' in sections:
        for item in sections['TOC']:
            key = (item.get('inputDate'), item.get('siteCd'), item.get('sampleCategory'))
            if key not in result_map:
                result_map[key] = {}

            toc_tc = item.get('tocTcMg')
            toc_ic = item.get('tocIcMg')
            if toc_tc:
                result_map[key]['TOC'] = str(toc_tc)
            elif toc_ic:
                result_map[key]['TOC'] = str(toc_ic)

    # Coli: coliA/coliB
    if 'Coli' in sections:
        for item in sections['Coli']:
            key = (item.get('inputDate'), item.get('siteCd'), item.get('sampleCategory'))
            if key not in result_map:
                result_map[key] = {}

            coli_a = item.get('coliA')
            coli_b = item.get('coliB')
            if coli_a or coli_b:
                result_map[key]['대장균군'] = f"{coli_a}/{coli_b}" if coli_a and coli_b else (coli_a or coli_b)

    return result_map

def convert_to_excel(txt_file, excel_file):
    """JSON 데이터를 Excel로 변환 (v2: 정확한 결과값 추출)"""

    print(f"txt 파일 읽는 중: {txt_file}")
    sections = extract_data_by_section(txt_file)
    result_map = build_result_map(sections)
    print(f"처리된 결과: {len(result_map)}개")

    # OCR 데이터를 기준으로 샘플 그룹화
    ocr_data = sections.get('OCR', [])
    by_facility = {}

    for record in ocr_data:
        siteCd = record.get('siteCd', '')

        # 마스터 시료 스킵
        if siteCd in ['SITE053', 'SITE065', 'SITE066']:
            continue

        facility = get_facility_name(siteCd)
        if facility not in by_facility:
            by_facility[facility] = []
        by_facility[facility].append(record)

    # Excel 생성
    wb = Workbook()
    wb.remove(wb.active)

    # 스타일 정의
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 시설별 시트 생성
    for facility, facility_records in sorted(by_facility.items()):
        ws = wb.create_sheet(title=facility)

        # 헤더
        headers = ['날짜', '시료명', '분류', 'BOD', 'TOC', 'SS', 'T-N', 'T-P', '대장균군', '입력자', '입력일시']
        ws.append(headers)

        # 헤더 포맷
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # 데이터 행 추가
        for record in sorted(facility_records, key=lambda x: x.get('inputDate', '')):
            input_date = record.get('inputDate', '')
            site_cd = record.get('siteCd', '')
            sample_category = record.get('sampleCategory', '')

            # 결과맵에서 결과값 조회
            key = (input_date, site_cd, sample_category)
            results = result_map.get(key, {})

            row = [
                input_date,
                record.get('sampleCategoryNm', ''),
                sample_category,
                results.get('BOD', ''),
                results.get('TOC', ''),
                results.get('SS', ''),
                results.get('T-N', ''),
                results.get('T-P', ''),
                results.get('대장균군', ''),
                record.get('createUser', ''),
                record.get('createTime', '')
            ]

            ws.append(row)

        # 열 너비 자동 조정
        for col_num, header in enumerate(headers, 1):
            max_length = len(header)
            column = get_column_letter(col_num)

            for row_num in range(2, ws.max_row + 1):
                cell_value = str(ws.cell(row=row_num, column=col_num).value or '')
                max_length = max(max_length, len(cell_value))

            ws.column_dimensions[column].width = min(max_length + 2, 30)

        # 데이터 셀 포맷
        for row_num in range(2, ws.max_row + 1):
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 저장
    wb.save(excel_file)
    print(f"✓ Excel 저장 완료: {excel_file}")

if __name__ == '__main__':
    txt_file = Path.home() / 'Documents/ETA/Data/facility_data_last_week.txt'
    excel_file = Path.home() / 'Documents/ETA/Data/facility_data_last_week_v2.xlsx'

    convert_to_excel(str(txt_file), str(excel_file))
