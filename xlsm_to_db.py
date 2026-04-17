#!/usr/bin/env python3
"""
청하 XLSM → 생물화학적_산소요구량_시험기록부 마이그레이션
BOD-DATA 시트의 데이터를 MariaDB에 직접 INSERT
"""

import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime, timedelta
import json
import sys

# MariaDB 연결
import mysql.connector

DB_CONFIG = {
    'host': '1110s.synology.me',
    'user': 'eta_user',
    'password': '1212xx!!AA',
    'database': 'eta_db',
    'charset': 'utf8mb4'
}

def excel_date_to_datetime(excel_date):
    """Excel 날짜 시리얼 → datetime 변환"""
    if isinstance(excel_date, str):
        # "2020. 1. 1. 오전 12:00:00" 형식 처리
        try:
            date_str = excel_date.split('. 오')[0]  # 시간 부분 제거
            date_str = date_str.replace('.', '-').replace(' ', '')
            return date_str
        except:
            return None
    elif isinstance(excel_date, (int, float)):
        # Excel 날짜 시리얼 (1900-02-29 버그 처리)
        if excel_date > 59:
            excel_date -= 1
        base_date = datetime(1899, 12, 30)
        return (base_date + timedelta(days=excel_date)).strftime('%Y-%m-%d')
    return None

def load_bod_data(xlsm_path):
    """BOD-DATA 시트에서 데이터 로드"""
    print(f"XLSM 파일 읽는 중: {xlsm_path}")
    wb = openpyxl.load_workbook(xlsm_path)

    # BOD-DATA 시트 찾기
    bod_sheet = None
    for sheet_name in wb.sheetnames:
        if sheet_name == 'BOD-DATA':
            bod_sheet = wb[sheet_name]
            print(f"✓ '{sheet_name}' 시트 발견")
            break

    if not bod_sheet:
        print("❌ BOD-DATA 시트를 찾을 수 없습니다")
        return []

    records = []

    # 데이터 범위: 2행~2252행
    for row_num in range(2, min(2253, bod_sheet.max_row + 1)):
        # A열: 날짜
        date_val = bod_sheet.cell(row_num, 1).value
        if not date_val:
            continue

        analysis_date = excel_date_to_datetime(date_val)
        if not analysis_date:
            continue

        # 고정 영역 (B~K 열)
        analyzer1 = bod_sheet.cell(row_num, 2).value or ''  # BOD-분석자
        analyzer2 = bod_sheet.cell(row_num, 3).value or ''  # 담당자-2
        seed_sample_ml = bod_sheet.cell(row_num, 4).value or ''  # 시료량
        do_15min = bod_sheet.cell(row_num, 5).value or ''  # 15min DO
        do_5day = bod_sheet.cell(row_num, 6).value or ''  # 5Day DO
        seed_bod = bod_sheet.cell(row_num, 7).value or ''  # 식종액 BOD
        seed_content = bod_sheet.cell(row_num, 8).value or ''  # 식종액함유량
        dilution_sample_ml = bod_sheet.cell(row_num, 9).value or ''  # 희석수 시료량
        seed_d1 = bod_sheet.cell(row_num, 10).value or ''  # 식종수 D1
        seed_d2 = bod_sheet.cell(row_num, 11).value or ''  # 식종수 D2

        # 샘플 블록 (col 30부터 8컬럼 간격)
        block_col = 30
        block_idx = 0

        while block_col <= bod_sheet.max_column:
            sample_name = bod_sheet.cell(row_num, block_col).value
            if not sample_name:
                break

            sample_ml = bod_sheet.cell(row_num, block_col + 1).value or ''
            d1 = bod_sheet.cell(row_num, block_col + 2).value or ''
            d2 = bod_sheet.cell(row_num, block_col + 3).value or ''
            dilution = bod_sheet.cell(row_num, block_col + 5).value or ''
            result = bod_sheet.cell(row_num, block_col + 6).value or ''
            sn = bod_sheet.cell(row_num, block_col + 7).value or ''

            # 결과값 없거나 0이면 skip
            if not result or result == '0':
                block_col += 8
                block_idx += 1
                continue

            # SN 기반 구분 판별
            division = '여수'
            if sn:
                sn_str = str(sn)
                if '율촌' in sn_str:
                    division = '율촌'
                    sn = sn_str.replace('[율촌]', '')
                elif '세풍' in sn_str:
                    division = '세풍'
                    sn = sn_str.replace('[세풍]', '')

            record = {
                '분석일': analysis_date,
                'SN': sn or '',
                '업체명': sample_name,
                '구분': division,
                '소스구분': '폐수배출업소',
                '시료명': sample_name,
                '시료량': sample_ml,
                'D1': d1,
                'D2': d2,
                '희석배수': dilution,
                '결과': result,
                '식종시료량': seed_sample_ml,
                '식종D1': seed_d1,
                '식종D2': seed_d2,
                '식종BOD': seed_bod,
                '식종함유량': seed_content,
                '비고': '',
                '등록일시': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            records.append(record)

            block_col += 8
            block_idx += 1

    print(f"✓ {len(records)}개 레코드 로드 완료")
    return records

def insert_to_db(records):
    """DB에 데이터 INSERT"""
    if not records:
        print("❌ 로드할 레코드가 없습니다")
        return False

    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()

        insert_query = """
            INSERT IGNORE INTO `생물화학적_산소요구량_시험기록부`
            (분석일, SN, 업체명, 구분, 소스구분, 시료명, 시료량, D1, D2,
             희석배수, 결과, 식종시료량, 식종D1, 식종D2, 식종BOD, 식종함유량,
             비고, 등록일시)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        inserted = 0
        for record in records:
            try:
                values = (
                    record['분석일'],
                    record['SN'],
                    record['업체명'],
                    record['구분'],
                    record['소스구분'],
                    record['시료명'],
                    record['시료량'],
                    record['D1'],
                    record['D2'],
                    record['희석배수'],
                    record['결과'],
                    record['식종시료량'],
                    record['식종D1'],
                    record['식종D2'],
                    record['식종BOD'],
                    record['식종함유량'],
                    record['비고'],
                    record['등록일시']
                )
                cursor.execute(insert_query, values)
                inserted += 1

                # 진행 표시
                if inserted % 50 == 0:
                    print(f"  {inserted}/{len(records)}...")
            except Exception as e:
                print(f"  ⚠ Row 에러: {e}")
                continue

        conn.commit()
        print(f"✓ {inserted}개 행 INSERT 완료")

        cursor.close()
        conn.close()
        return True

    except Exception as e:
        print(f"❌ DB 연결 오류: {e}")
        return False

if __name__ == '__main__':
    xlsm_file = Path.home() / 'Documents/ETA/Docs/CHUNGHA-김가린.xlsm'

    # 1. XLSM 파일 읽기
    records = load_bod_data(str(xlsm_file))

    if records:
        # 2. DB에 INSERT
        if insert_to_db(records):
            print("\n✅ 마이그레이션 완료!")
        else:
            print("\n❌ 마이그레이션 실패")
            sys.exit(1)
    else:
        print("\n❌ 로드할 데이터 없음")
        sys.exit(1)
